from multiprocessing import Value
from re import S
import openpyxl

inv_file = openpyxl.load_workbook("spreadsheet/inventory.xlsx")
product_list = inv_file["Sheet1"]

prod_per_supply = {}
total_value_per_supply = {}
prod_num_count = {}
for prod_row in range(2, product_list.max_row + 1):
    supply_name = product_list.cell(prod_row, 4).value
    inventory = product_list.cell(prod_row, 2).value
    price = product_list.cell(prod_row, 3).value
    prod_num = product_list.cell(prod_row, 1).value
    inventory_price = product_list.cell(prod_row, 5)

    # Calculating supplier 
    if supply_name in prod_per_supply:
        prod_per_supply[supply_name] = prod_per_supply[supply_name]+1
    
    else:
        prod_per_supply[supply_name] = 1

    # Calculating the Value of supply
    if supply_name in total_value_per_supply:
        total_value_per_supply[supply_name] = total_value_per_supply[supply_name] + inventory * price
    
    else:
        total_value_per_supply[supply_name] = inventory * price
    

    # Printing the product number less than 10 inventory
    if inventory < 10 :
        prod_num_count[int(prod_num)] = int(inventory)

    # Adding data to a sheet
    inventory_price.value = inventory * price

inv_file.save("spreadsheet/inventory_with_total.xlsx")

print(prod_per_supply)
print(total_value_per_supply)
print(prod_num_count)